"""
============================================================
SmartTracer ETL v5: Excel Bersih (2019-2024) → PostgreSQL OLTP
============================================================
PERUBAHAN dari v4:
  1. FIX f502 DUPLIKAT: f502 dipastikan SELALU answer_index = 0
     (sebelumnya ada celah karena 2 keyword regex yang sama-sama
     match f502 di kolom Excel berbeda → menumpuk index 0,1,2,dst
     di beberapa file/tahun yang strukturnya berbeda)
  2. NULL/KOSONG → diisi "0" (bukan di-skip lagi).
     Berlaku untuk semua question_code KECUALI provinsi/kota
     (f5a1/f5a2 tetap NULL kalau kosong, karena 0 bukan id valid)
  3. CLEANUP OTOMATIS: sebelum insert ulang, ETL menghapus dulu
     answer_index > 0 yang nyasar untuk fcode non-boolean
     (safety net kalau ada sisa data lama yang menumpuk)
============================================================
"""

import os, sys, re, traceback, json
import psycopg2
import openpyxl
from datetime import datetime, date

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "tracer_oltp",
    "user":     "postgres",
    "password": "Rin_2722",
}

EXCEL_FILES = [
    ("data/Data_2019.xlsx", 2019),
    ("data/Data_2020.xlsx", 2020),
    ("data/Data_2021.xlsx", 2021),
    ("data/Data_2022.xlsx", 2022),
    ("data/Data_2023.xlsx", 2023),
    ("data/Data_2024.xlsx", 2024),
]

RUN_ALL = True   # False = testing satu file saja
# TEST_FILE       = "data/Data_2024.xlsx"
# TEST_GRAD_YEAR  = 2024
# TEST_MAX_ROWS   = 50

# ══════════════════════════════════════════════════════════════════
# SHEET → PROGRAM CODE
# ══════════════════════════════════════════════════════════════════
SHEET_TO_PROGRAM_CODE = {
    "D3 - Teknik Konstruksi Gedung":         "TKG",
    "D3 - Teknik Konstruksi Sipil":          "TKS",
    "D4 - Teknik Perancangan Jalan Dan":     "TPJJ",
    "D4 - Teknik Perancangan Jalan d":       "TPJJ",
    "D4 - Teknik Perawatan Dan Perba":       "TPPG",
    "D3 - Teknik Mesin":                     "TM",
    "D4 - Teknik Perancangan dan Kon":       "TPKM",
    "D4 - Proses Manufaktur":                "PM",
    "D3 - Teknik Pendingin Dan Tata ":       "TPTU3",
    "D4 - Teknik Pendingin dan Tata ":       "TPTU4",
    "D4 - Teknik Pendingin dan Tata":        "TPTU4",
    "D3 - Teknik Konversi Energi":           "TKE3",
    "D4 - Teknik Konservasi Energi":         "TKE4",
    "D4 - Teknologi Pembangkit Tenag":       "TPTL",
    "D3 - Teknik Elektronika":               "TEL3",
    "D4 - Teknik Elektronika":               "TEL4",
    "D3 - Teknik Listrik":                   "TL",
    "D4 - Teknik Otomasi Industri":          "TOI",
    "D3 - Teknik Telekomunikasi":            "TELKOM3",
    "D4 - Teknik Telekomunikasi":            "TELKOM4",
    "D3 - Teknik Kimia":                     "TK3",
    "D3 - Analis Kimia":                     "AK3",
    "D4 - Teknik Kimia Produksi Bers":       "TKPB",
    "D3 - Teknik Informatika":               "TI3",
    "D4 - Teknik Informatika":               "TI",
    "D3 - Teknik Aeronautika":               "TA",
    "D3 - Akuntansi":                        "AKT3",
    "D4 - Akuntansi":                        "AKT4",
    "D3 - Keuangan Dan Perbankan":           "KP",
    "D4 - Keuangan Syariah":                 "KS",
    "D4 - Akuntansi Manajemen Pemeri":       "AMP",
    "D4 - Akuntansi Manajemen Pemerintahan": "AMP",
    "D4 - Manajemen Aset":                   "MA",
    "D3 - Administrasi Bisnis":              "AB3",
    "D4 - Administrasi Bisnis":              "AB4",
    "D3 - Manajemen Pemasaran":              "MP3",
    "D4 - Manajemen Pemasaran":              "MP4",
    "D3 - Usaha Perjalanan Wisata":          "UPW",
    "D3 - Bahasa Inggris":                   "BIG",
    "S2 - Keuangan dan Perbankan Sya":       "KPS2",
    "S2 - Rekayasa Infrastruktur (Te":       "RIS2",
    "S2 - Keuangan dan Perbankan Syariah":   "KPS2",
    "S2 - Rekayasa Infrastruktur":           "RIS2",
    "Sheet1": None,
}

# ══════════════════════════════════════════════════════════════════
# STATUS KERJA → label PDDIKTI
# ══════════════════════════════════════════════════════════════════
STATUS_TO_LABEL = {
    1: "Bekerja (full time / part time)",
    2: "Belum memungkinkan bekerja",
    3: "Wiraswasta",
    4: "Melanjutkan Pendidikan",
    5: "Tidak kerja tetapi sedang mencari kerja",
    6: "Melanjutkan pendidikan sambil bekerja",
    7: "Melanjutkan pendidikan sambil wiraswasta",
}

STATUS_BEKERJA   = {1, 6}
STATUS_WIRAUSAHA = {3, 7}
STATUS_ADA_KERJA = STATUS_BEKERJA | STATUS_WIRAUSAHA

# ══════════════════════════════════════════════════════════════════
# MAPPING PROVINSI (fallback)
# ══════════════════════════════════════════════════════════════════
PROVINCE_NAME_TO_CODE = {
    "DKI Jakarta":"10000","Jawa Barat":"20000","Jawa Tengah":"30000",
    "DI Yogyakarta":"40000","Jawa Timur":"50000","Banten":"280000",
    "Aceh":"60000","Sumatera Utara":"70000","Sumatera Barat":"80000",
    "Riau":"90000","Jambi":"100000","Sumatera Selatan":"110000",
    "Lampung":"120000","Kepulauan Bangka Belitung":"290000",
    "Bengkulu":"260000","Kepulauan Riau":"310000",
    "Kalimantan Barat":"130000","Kalimantan Tengah":"140000",
    "Kalimantan Selatan":"150000","Kalimantan Timur":"160000",
    "Kalimantan Utara":"340000","Sulawesi Utara":"170000",
    "Sulawesi Tengah":"180000","Sulawesi Selatan":"190000",
    "Sulawesi Tenggara":"200000","Gorontalo":"300000",
    "Sulawesi Barat":"330000","Bali":"220000",
    "Nusa Tenggara Barat":"230000","Nusa Tenggara Timur":"240000",
    "Maluku":"210000","Maluku Utara":"270000",
    "Papua":"250000","Papua Barat":"320000",
    "Luar Negeri":"350000",
}
PROVINCE_INDEX_TO_CODE = {
    1:"60000",2:"70000",3:"80000",4:"90000",5:"310000",
    6:"100000",7:"110000",8:"290000",9:"260000",10:"120000",
    11:"10000",12:"280000",13:"20000",14:"30000",15:"40000",
    16:"50000",17:"220000",18:"230000",19:"240000",20:"130000",
    21:"140000",22:"150000",23:"160000",24:"340000",25:"170000",
    26:"300000",27:"180000",28:"330000",29:"190000",30:"200000",
    31:"210000",32:"270000",33:"320000",34:"250000",35:"350000",
}

CITY_PREFIX_TO_PROV = {
    "jabar":"20000","jawa barat":"20000",
    "dki jakarta":"10000","dki":"10000",
    "jateng":"30000","jawa tengah":"30000",
    "diy":"40000","di yogyakarta":"40000","yogyakarta":"40000",
    "jatim":"50000","jawa timur":"50000",
    "banten":"280000","aceh":"60000",
    "sumatera utara":"70000","sumut":"70000",
    "sumatera barat":"80000","sumbar":"80000",
    "riau":"90000","jambi":"100000",
    "sumatera selatan":"110000","sumsel":"110000",
    "lampung":"120000",
    "bangka belitung":"290000","babel":"290000",
    "bengkulu":"260000",
    "kepulauan riau":"310000","kepri":"310000","kep. riau":"310000",
    "kalimantan barat":"130000","kalbar":"130000",
    "kalimantan tengah":"140000","kalteng":"140000",
    "kalimantan selatan":"150000","kalsel":"150000",
    "kalimantan timur":"160000","kaltim":"160000",
    "kalimantan utara":"340000","kaltara":"340000",
    "sulawesi utara":"170000","sulut":"170000",
    "sulawesi tengah":"180000","sulteng":"180000",
    "sulawesi selatan":"190000","sulsel":"190000",
    "sulawesi tenggara":"200000","sultra":"200000",
    "gorontalo":"300000",
    "sulawesi barat":"330000","sulbar":"330000",
    "bali":"220000",
    "nusa tenggara barat":"230000","ntb":"230000",
    "nusa tenggara timur":"240000","ntt":"240000",
    "maluku":"210000","maluku utara":"270000",
    "papua":"250000","papua barat":"320000",
    "luar negeri":"350000",
}

# ══════════════════════════════════════════════════════════════════
# LOOKUP GEO (diisi dari DB saat startup)
# ══════════════════════════════════════════════════════════════════
DB_PROVINCES = {}
DB_CITIES    = {}

def load_geo(cur):
    cur.execute("SELECT id, code FROM tracer_oltp.provinces")
    for pid, pcode in cur.fetchall():
        DB_PROVINCES[pcode] = pid
    cur.execute("SELECT id, province_code, name FROM tracer_oltp.cities")
    for cid, pcode, cname in cur.fetchall():
        DB_CITIES.setdefault(pcode, {})[cname.lower()] = (cid, cname)
    print(f"  Loaded {len(DB_PROVINCES)} provinsi, "
          f"{sum(len(v) for v in DB_CITIES.values())} kota dari DB")

# ══════════════════════════════════════════════════════════════════
# MAPPING KOLOM EXCEL → question_code
# ══════════════════════════════════════════════════════════════════
COL_KEYWORD_TO_FCODE = {
    # ── Identitas kementrian ─────────────────────────────────────
    r"Kode PT":                                   "kdptimsmh",
    r"Tahun Lulus":                               "tahun_lulus",
    r"Kode Prodi":                                "kdpstmsmh",
    r"Nomor Telepon":                             "telpomsmh",
    r"Alamat Email":                              "emailmsmh",
    r"\bNIK\b":                                   "nik",
    r"\bNPWP\b":                                  "npwp",

    # ── Status & pekerjaan ───────────────────────────────────────
    r"status Anda saat ini":                      "f8",
    r"STATUS ANDA SAAT INI":                      "f8",

    # f502: dipecah jadi 2 fcode SEMENTARA saat deteksi kolom —
    # f502 (general, PRIORITAS) dan f502_alt (pekerjaan pertama / wiraswasta,
    # FALLBACK). Saat insert, kalau f502 general ADA ISI → itu yang dipakai.
    # Kalau f502 general kosong tapi f502_alt ada isi → pakai f502_alt.
    # Ini menangani 5 kasus di 2021 dimana DUA kolom sama-sama terisi
    # dengan nilai berbeda — sesuai instruksi: selalu pilih kolom
    # "berapa bulan mendapatkan pekerjaan" (general), bukan "pekerjaan pertama".
    r"berapa bulan.*mendapatkan pekerjaan\s*\?":   "f502",        # general — PRIORITAS
    r"berapa bulan.*mendapatkan pekerjaan pertama": "f502_alt",   # fallback
    r"berapa bulan.*memulai wiraswasta":            "f502_alt",   # fallback
    # CATATAN: keyword "mendapatkan pekerjaan <= 6 bulan" DIHAPUS dari sini.
    # Itu pertanyaan ya/tidak ("Apakah anda telah mendapatkan pekerjaan <= 6
    # bulan / termasuk bekerja sebelum lulus?"), BUKAN pertanyaan jumlah bulan.
    # Kalau di-map ke f502, nilainya (1/2) salah tertimpa ke jawaban f502 yang
    # asli (jumlah bulan). Root cause bug 5 alumni 2021 yang semua jadi "1".

    r"rata-rata pendapatan.*per bulan":            "f505",
    r"rata-rata pendapatan.*take home":            "f505",
    r"pendapatan anda saat ini":                   "f505",

    # Lokasi kerja — PROVINSI → answer_text = id provinsi (integer)
    r"lokasi.*bekerja.*Provinsi":                  "f5a1",
    r"lokasi.*bekerja.*Propinsi":                  "f5a1",
    r"Propinsi.*bekerja":                          "f5a1",

    # Lokasi kerja — KOTA/KABUPATEN → answer_text = id kota (integer)
    r"lokasi.*bekerja.*Kota":                      "f5a2",
    r"lokasi.*bekerja.*Kabupaten":                 "f5a2",
    r"Kabupaten.*Kota.*bekerja":                   "f5a2",
    r"Kota.*Kabupaten.*bekerja":                   "f5a2",

    r"jenis perusahaan.*instansi.*institusi.*tempat": "f1101",
    r"jenis perusahaan.*instansi lainnya":          "f1102",
    r"nama perusahaan.*kantor":                     "f5b",
    r"posisi.*jabatan.*wiraswasta":                 "f5c",
    r"tingkat.*tempat kerja":                       "f5d",
    r"kelompok.*tingkat.*tempat kerja":             "f5d",

    # ── Studi lanjut ─────────────────────────────────────────────
    r"sumber.*biaya.*studi lanjut":                 "f18a",
    r"sumber beasiswa studi lanjut":                "f18a",
    r"Perguruan Tinggi.*studi lanjut":              "f18b",
    r"Nama Perguruan Tinggi.*melanjutkan":          "f18b",
    r"Program Studi.*studi lanjut":                 "f18c",
    r"Jenis Program Studi":                         "f18c",
    r"Tanggal Masuk.*studi":                        "f18d",

    # ── Sumber dana kuliah ───────────────────────────────────────
    r"sumberdana.*pembiayaan kuliah":               "f1201",
    r"sumber dana.*pembiayaan kuliah":              "f1201",
    r"sumber dana pembiayaan.*lainnya":             "f1202",

    # ── Relevansi & kecocokan pendidikan ─────────────────────────
    r"erat.*hubungan.*bidang studi":                "f14",
    r"keterkaitan.*bidang studi":                   "f14",
    r"paling tepat.*sesuai.*pekerjaan":             "f15",

    # ── Kompetensi saat lulus (A) ─────────────────────────────────
    r"saat LULUS.*Etika":                           "f1761",
    r"saat LULUS.*Keahlian berdasarkan":            "f1763",
    r"saat LULUS.*Bahasa Inggris":                  "f1765",
    r"saat LULUS.*Penggunaan Teknologi":            "f1767",
    r"saat LULUS.*Komunikasi":                      "f1769",
    r"saat LULUS.*Kerja sama tim":                  "f1771",
    r"saat LULUS.*Pengembangan":                    "f1773",

    # ── Kompetensi saat ini / di pekerjaan (B) ───────────────────
    r"saat ini.*Etika":                             "f1762",
    r"saat ini.*Keahlian berdasarkan":              "f1764",
    r"saat ini.*Bahasa Inggris":                    "f1766",
    r"saat ini.*Penggunaan Teknologi":              "f1768",
    r"saat ini.*Komunikasi":                        "f1770",
    r"saat ini.*Kerja sama tim":                    "f1772",
    r"saat ini.*Pengembangan":                      "f1774",

    # ── Metode pembelajaran ──────────────────────────────────────
    r"penekanan.*\[Perkuliahan dalam Prodi\]":   "f21",
    r"penekanan.*\[Magang\]":                    "f24",
    r"penekanan.*\[Praktikum":                   "f25",
    r"penekanan.*\[Kerja Lapangan\]":            "f26",
    r"penekanan.*\[Diskusi\]":                   "f27",
    r"penekanan.*Perkuliahan(?!.*luar)":         "f21",
    r"penekanan.*Magang":                        "f24",
    r"penekanan.*Praktikum":                     "f25",
    r"penekanan.*Kerja Lapangan":                "f26",
    r"penekanan.*Diskusi":                       "f27",
    r"penekanan.*Demonstrasi":                   "f22",
    r"penekanan.*Partisipasi":                   "f23",
    r"penekanan.*[Pp]royek [Rr]iset":             "f23",
    # f22, f23 → tidak ada padanan di Excel POLBAN, memang tidak terisi.

    # ── Kapan mulai mencari kerja ────────────────────────────────
    r"Kapan.*mulai mencari pekerjaan":              "f301",
    r"berapa bulan.*sebelum lulus.*mencari":        "f302",
    r"berapa bulan.*sesudah lulus.*mencari":        "f303",
    r"SEBELUM LULUS.*mencari":                      "f302",
    r"SETELAH [Ll]ulus.*mencari":                   "f303",

    # ── Cara mencari kerja (f401-f415, boolean 0/1) ──────────────
    r"mencari pekerjaan.*\[1\]":                    "f401",
    r"mencari pekerjaan.*\[2\]":                    "f402",
    r"mencari pekerjaan.*\[3\]":                    "f403",
    r"mencari pekerjaan.*\[4\]":                    "f404",
    r"mencari pekerjaan.*\[5\]":                    "f405",
    r"mencari pekerjaan.*\[6\]":                    "f406",
    r"mencari pekerjaan.*\[7\]":                    "f407",
    r"mencari pekerjaan.*\[8\]":                    "f408",
    r"mencari pekerjaan.*\[9\]":                    "f409",
    r"mencari pekerjaan.*\[10\]":                   "f410",
    r"mencari pekerjaan.*\[11\]":                   "f411",
    r"mencari pekerjaan.*\[12\]":                   "f412",
    r"mencari pekerjaan.*\[13\]":                   "f413",
    r"mencari pekerjaan.*\[14\]":                   "f414",
    r"mencari pekerjaan.*\[other\]":                "f415",
    r"cara lainnya.*mencari pekerjaan":             "f416",

    # ── Jumlah lamaran ───────────────────────────────────────────
    r"sudah Anda lamar.*surat.*e-mail":             "f6",
    r"sudah anda lamar.*surat":                     "f6",
    r"yang merespons lamaran":                      "f7",
    r"yang mengundang.*wawancara":                  "f7a",

    # ── Aktif mencari kerja 4 minggu ─────────────────────────────
    r"aktif mencari pekerjaan.*4 minggu":           "f1001",
    r"aktivitas lainnya.*mencari pekerjaan":        "f1002",

    # ── Alasan ketidaksesuaian (f1601-f1613, boolean) ────────────
    r"mengapa.*mengambilnya.*\[1\]":                "f1601",
    r"mengapa.*mengambilnya.*\[2\]":                "f1602",
    r"mengapa.*mengambilnya.*\[3\]":                "f1603",
    r"mengapa.*mengambilnya.*\[4\]":                "f1604",
    r"mengapa.*mengambilnya.*\[5\]":                "f1605",
    r"mengapa.*mengambilnya.*\[6\]":                "f1606",
    r"mengapa.*mengambilnya.*\[7\]":                "f1607",
    r"mengapa.*mengambilnya.*\[8\]":                "f1608",
    r"mengapa.*mengambilnya.*\[9\]":                "f1609",
    r"mengapa.*mengambilnya.*\[10\]":               "f1610",
    r"mengapa.*mengambilnya.*\[11\]":               "f1611",
    r"mengapa.*mengambilnya.*\[12\]":               "f1612",
    r"mengapa.*mengambilnya.*\[other\]":            "f1613",
    r"alasan lainnya.*tidak sesuai":                "f1614",

    # ── Misc ─────────────────────────────────────────────────────
    r"jumlah sertifikasi":                          "f_sertif",
}

# question_code yang berisi nilai provinsi (perlu di-resolve ke province_id)
FCODE_PROVINCE = {"f5a1"}
# question_code yang berisi nilai kota (perlu di-resolve ke city_id)
FCODE_CITY     = {"f5a2"}

# fcode boolean (multi-pilihan): boleh banyak kolom, tiap kolom = answer_index unik
BOOLEAN_FCODES = {
    'f401','f402','f403','f404','f405','f406','f407','f408','f409','f410',
    'f411','f412','f413','f414','f415',
    'f1601','f1602','f1603','f1604','f1605','f1606','f1607','f1608',
    'f1609','f1610','f1611','f1612','f1613',
}

# fcode yang TIDAK boleh diisi "0" kalau kosong (0 bukan ID valid untuk geo)
FCODE_NO_ZERO_FILL = {"f5a1", "f5a2"}

# question_code identitas alumni (disimpan juga ke alumni_profiles)
FCODE_IDENTITY = {
    "nimhsmsmh", "kdptimsmh", "tahun_lulus", "kdpstmsmh",
    "nmmhsmsmh", "telpomsmh", "emailmsmh", "nik", "npwp",
}

# ══════════════════════════════════════════════════════════════════
# GEO HELPERS
# ══════════════════════════════════════════════════════════════════

def resolve_province_id(raw):
    if raw is None: return None
    s = str(raw).strip()
    if s in ('', 'nan', 'None', '-'): return None
    try:
        n = int(float(s))
        if 1 <= n <= 40:
            pcode = PROVINCE_INDEX_TO_CODE.get(n)
            return str(DB_PROVINCES[pcode]) if pcode and pcode in DB_PROVINCES else None
        s_code = str(n)
        if s_code in DB_PROVINCES:
            return str(DB_PROVINCES[s_code])
    except ValueError:
        pass
    pcode = PROVINCE_NAME_TO_CODE.get(s)
    if pcode and pcode in DB_PROVINCES:
        return str(DB_PROVINCES[pcode])
    parts = re.split(r'\s*[-–]\s*', s, maxsplit=1)
    if len(parts) == 2:
        prefix_pcode = CITY_PREFIX_TO_PROV.get(parts[0].strip().lower())
        if prefix_pcode and prefix_pcode in DB_PROVINCES:
            return str(DB_PROVINCES[prefix_pcode])
    return None


def resolve_city_id(raw, province_id_str):
    if raw is None: return None
    s = str(raw).strip()
    if s in ('', 'nan', 'None', '-'): return None
    pcode = None
    if province_id_str:
        try:
            pid_int = int(province_id_str)
            for pc, pid in DB_PROVINCES.items():
                if pid == pid_int:
                    pcode = pc
                    break
        except: pass
    parts = re.split(r'\s*[-–]\s*', s, maxsplit=1)
    city_part = parts[1].strip() if len(parts) == 2 else s.strip()
    if not pcode and len(parts) == 2:
        prefix_pcode = CITY_PREFIX_TO_PROV.get(parts[0].strip().lower())
        if prefix_pcode:
            pcode = prefix_pcode
    if not pcode:
        return None
    prov_cities = DB_CITIES.get(pcode, {})
    city_norm = re.sub(r'^kota\s+administrasi\s+', 'Kota ', city_part, flags=re.I)
    city_norm = re.sub(r'^kabupaten\s+', 'Kab. ', city_norm, flags=re.I)
    entry = prov_cities.get(city_norm.lower())
    if entry:
        return str(entry[0])
    clean = re.sub(r'^(kab\.\s*|kota\s*)', '', city_norm.lower()).strip()
    for dbl, (dbi, _) in prov_cities.items():
        clean_db = re.sub(r'^(kab\.\s*|kota\s*)', '', dbl).strip()
        if clean == clean_db:
            return str(dbi)
    return None

# ══════════════════════════════════════════════════════════════════
# HELPERS UMUM
# ══════════════════════════════════════════════════════════════════

def is_null(v):
    if v is None: return True
    return str(v).strip() in ('', 'nan', 'None', 'NULL', '-', '.')
    # CATATAN: '0' SENGAJA tidak dianggap null lagi di v5 — kalau Excel
    # memang berisi angka 0, itu jawaban valid, bukan kekosongan.

def clean_nim(v):
    if v is None: return None
    s = str(v).strip()
    if s in ('', 'nan', 'None'): return None
    return s.split('.')[0] if '.' in s else s

def clean_salary(v):
    if v is None: return None
    s = re.sub(r'[Rr][Pp]\.?\s*', '', str(v).strip())
    if s in ('', 'nan', 'None', '-'): return None
    if s.count('.') > 1:
        s = s.replace('.', '')
    elif s.count('.') == 1 and len(s.split('.')[1]) == 3:
        s = s.replace('.', '')
    s = s.replace(',', '.')
    try:
        n = float(s)
        if n < 0: return None
        return n * 1_000_000 if 0 < n <= 100 else n
    except:
        return None

def get_program_id(sheet_name, cur):
    code = SHEET_TO_PROGRAM_CODE.get(sheet_name)
    if code is None:
        for key, val in SHEET_TO_PROGRAM_CODE.items():
            if key and (sheet_name.startswith(key) or key.startswith(sheet_name[:20])):
                code = val
                break
    if not code:
        return None
    cur.execute("SELECT id FROM tracer_oltp.programs WHERE code = %s LIMIT 1", (code,))
    row = cur.fetchone()
    return row[0] if row else None

def get_questionnaire_id(grad_year, cur):
    grad_year_json = json.dumps([grad_year])
    cur.execute("""
        SELECT id FROM tracer_oltp.questionnaires
        WHERE code LIKE 'DIKTI_%%'
          AND target_graduation_years @> %s::jsonb
        LIMIT 1
    """, (grad_year_json,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("""
        SELECT id FROM tracer_oltp.questionnaires
        WHERE code LIKE 'DIKTI_%%'
        ORDER BY id ASC
        LIMIT 1
    """)
    row = cur.fetchone()
    return row[0] if row else 1

def build_col_map(headers):
    """
    Scan headers Excel, cocokkan ke COL_KEYWORD_TO_FCODE.
    Return dict: fcode → [col_index, ...]

    PENTING: untuk fcode NON-boolean, kalau ada lebih dari satu kolom
    yang match (karena header berbeda tapi sama-sama lolos regex),
    HANYA kolom pertama yang dipertahankan di sini juga (selain nanti
    di-enforce lagi saat insert) — supaya tidak ada ambiguitas index.
    """
    col_map = {}
    for ci, h in enumerate(headers):
        if not h: continue
        hs = str(h)
        for kw, fc in COL_KEYWORD_TO_FCODE.items():
            try:
                matched = bool(re.search(kw, hs, re.I | re.DOTALL))
            except:
                matched = kw.lower() in hs.lower()
            if matched:
                col_map.setdefault(fc, []).append(ci)
                break   # satu kolom → satu fcode, stop di match pertama
    return col_map

def get_cell(row, indices):
    for ci in (indices or []):
        if ci < len(row) and not is_null(row[ci]):
            return row[ci]
    return None

# ══════════════════════════════════════════════════════════════════
# INSERT response_answers
# ══════════════════════════════════════════════════════════════════

def upsert_answer(cur, rid, question_code, answer_text, answer_index=0):
    """
    Insert atau update satu jawaban ke response_answers.
    v5: kalau answer_text kosong/None DAN fcode bukan f5a1/f5a2,
    isi dengan "0" (bukan di-skip).
    f5a1/f5a2 tetap di-skip kalau kosong karena 0 bukan ID valid.
    """
    val = None if answer_text is None else str(answer_text).strip()
    if val in (None, '', 'nan', 'None'):
        if question_code in FCODE_NO_ZERO_FILL:
            return   # f5a1/f5a2 kosong → tidak diisi sama sekali
        val = "0"
    cur.execute("""
        INSERT INTO tracer_oltp.response_answers
            (response_id, question_code, answer_index, answer_text, created_at, updated_at)
        VALUES (%s, %s, %s, %s, NOW(), NOW())
        ON CONFLICT (response_id, question_code, answer_index) DO UPDATE SET
            answer_text = EXCLUDED.answer_text,
            updated_at  = NOW()
    """, (rid, question_code, answer_index, val))


def cleanup_stray_answers(cur, rid, fc):
    """
    Safety net: untuk fcode NON-boolean, hapus answer_index > 0 yang
    mungkin tertinggal dari run ETL sebelumnya (sebelum fix v5) —
    supaya tidak ada duplikat menumpuk per response_id.
    """
    if fc in BOOLEAN_FCODES:
        return
    cur.execute("""
        DELETE FROM tracer_oltp.response_answers
        WHERE response_id = %s AND question_code = %s AND answer_index > 0
    """, (rid, fc))

# ══════════════════════════════════════════════════════════════════
# ETL UTAMA
# ══════════════════════════════════════════════════════════════════

def process_file(filepath, grad_year, conn, max_rows=None):
    cur = conn.cursor()
    wb  = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ok  = err = 0

    for sname in wb.sheetnames:
        if sname == "Sheet1":
            continue

        pid = get_program_id(sname, cur)
        if pid is None:
            print(f"  [SKIP] {sname}")
            continue

        qid = get_questionnaire_id(grad_year, cur)
        ws  = wb[sname]
        it  = ws.iter_rows(values_only=True)
        headers  = list(next(it))
        col_map  = build_col_map(headers)
        s_ok = s_err = 0
        n_detected = len(col_map)
        print(f"  Processing: {sname} ({n_detected} kolom terdeteksi) ...")

        for rnum, row in enumerate(it, 2):
            if max_rows and rnum > max_rows + 1:
                break

            if not row or is_null(row[0] if row else None) or is_null(row[1] if len(row) > 1 else None):
                continue

            nama = re.sub(r'^[\d]+\.\s*', '', str(row[0]).strip()).strip()
            nim  = clean_nim(row[1])
            if not nim:
                continue

            try:
                # ── 1. alumni_profiles ────────────────────────────────────
                email_raw = get_cell(row, col_map.get("emailmsmh", []))
                phone_raw = get_cell(row, col_map.get("telpomsmh", []))
                nik_raw   = get_cell(row, col_map.get("nik", []))
                npwp_raw  = get_cell(row, col_map.get("npwp", []))
                kdpt_raw  = get_cell(row, col_map.get("kdptimsmh", []))

                email = str(email_raw).strip() if not is_null(email_raw) else None
                phone = str(phone_raw).strip() if not is_null(phone_raw) else None
                nik   = str(nik_raw).strip()   if not is_null(nik_raw)   else None
                npwp  = str(npwp_raw).strip()  if not is_null(npwp_raw)  else None
                kode_pt = str(kdpt_raw).strip() if not is_null(kdpt_raw) else None

                cur.execute("""
                    INSERT INTO tracer_oltp.alumni_profiles
                        (nim, name, email, phone, program_id, graduation_year,
                         nik, npwp, kode_pt, is_active, created_at, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,true,NOW(),NOW())
                    ON CONFLICT (nim) DO UPDATE SET
                        name            = EXCLUDED.name,
                        email           = COALESCE(EXCLUDED.email, tracer_oltp.alumni_profiles.email),
                        phone           = COALESCE(EXCLUDED.phone, tracer_oltp.alumni_profiles.phone),
                        nik             = COALESCE(EXCLUDED.nik,   tracer_oltp.alumni_profiles.nik),
                        npwp            = COALESCE(EXCLUDED.npwp,  tracer_oltp.alumni_profiles.npwp),
                        kode_pt         = COALESCE(EXCLUDED.kode_pt, tracer_oltp.alumni_profiles.kode_pt),
                        graduation_year = EXCLUDED.graduation_year,
                        updated_at      = NOW()
                    RETURNING id
                """, (nim, nama, email, phone, pid, grad_year, nik, npwp, kode_pt))
                aid = cur.fetchone()[0]

                # ── 2. responses ──────────────────────────────────────────
                # Satu alumni = satu response, walaupun muncul di beberapa file Excel
                cur.execute("""
                    SELECT id FROM tracer_oltp.responses
                    WHERE alumni_id = %s
                    ORDER BY id ASC LIMIT 1
                """, (aid,))
                existing = cur.fetchone()
                if existing:
                    rid = existing[0]
                    cur.execute("""
                        UPDATE tracer_oltp.responses
                        SET questionnaire_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (qid, rid))
                else:
                    cur.execute("""
                        INSERT INTO tracer_oltp.responses
                            (questionnaire_id, alumni_id, status, submitted_at, source, created_at, updated_at)
                        VALUES (%s,%s,'submitted',NOW(),'etl',NOW(),NOW())
                        RETURNING id
                    """, (qid, aid))
                    rid = cur.fetchone()[0]

                # ── 3. response_answers — SEMUA kolom Excel ───────────────
                upsert_answer(cur, rid, "nmmhsmsmh", nama, 0)
                upsert_answer(cur, rid, "nimhsmsmh", nim,  0)

                province_id_str = None

                # ── f502 khusus: gabungkan f502 (general) + f502_alt (pekerjaan
                # pertama/wiraswasta) SEBELUM loop utama. Prioritas: f502 general
                # menang kalau dua-duanya terisi (sesuai instruksi tim) — ini
                # menangani 5 kasus alumni 2021 yang kedua kolom sama-sama ada isi.
                if "f502" in col_map or "f502_alt" in col_map:
                    raw_general = get_cell(row, col_map.get("f502", []))
                    raw_alt     = get_cell(row, col_map.get("f502_alt", []))
                    # Prioritas: general dulu, fallback ke alt kalau general kosong
                    final_f502 = raw_general if not is_null(raw_general) else raw_alt
                    cleanup_stray_answers(cur, rid, "f502")
                    if is_null(final_f502):
                        upsert_answer(cur, rid, "f502", None, 0)
                    else:
                        upsert_answer(cur, rid, "f502", str(final_f502).strip(), 0)

                for fc, col_indices in col_map.items():
                    if fc in ("nimhsmsmh", "nmmhsmsmh", "f502", "f502_alt"):
                        # f502/f502_alt sudah ditangani khusus di atas
                        continue

                    # Cleanup safety net: hapus sisa answer_index > 0 dari
                    # run lama untuk fcode non-boolean
                    cleanup_stray_answers(cur, rid, fc)

                    if fc in BOOLEAN_FCODES:
                        # Boolean: semua kolom dipakai, tiap kolom = answer_index berbeda
                        # (tidak ada masalah "kolom mana yang terisi" karena ini
                        #  multi-pilihan, tiap kolom representasi pilihan terpisah)
                        for idx, ci in enumerate(col_indices):
                            if ci >= len(row):
                                continue
                            raw = row[ci]
                            if is_null(raw):
                                upsert_answer(cur, rid, fc, None, idx)
                            else:
                                upsert_answer(cur, rid, fc, str(raw).strip(), idx)
                        continue

                    # ── Non-boolean (termasuk f502) ──────────────────────
                    # PENTING: kalau ada beberapa kolom yang match fcode yang
                    # sama (misal "...pekerjaan pertama?" DAN "...memulai
                    # wiraswasta?" sama-sama → f502), kita AMBIL KOLOM YANG
                    # ADA ISINYA, bukan sekadar kolom pertama dalam urutan.
                    # Karena satu alumni cuma punya SATU dari dua kondisi itu
                    # (bekerja ATAU wiraswasta), get_cell() otomatis pilih
                    # mana yang terisi. answer_index SELALU 0 — tidak pernah
                    # lebih dari satu baris per (response_id, question_code).
                    raw = get_cell(row, col_indices)
                    idx = 0

                    if fc in FCODE_PROVINCE:
                        if is_null(raw):
                            continue   # f5a1 kosong → skip total, tidak isi 0
                        resolved = resolve_province_id(raw)
                        if resolved:
                            province_id_str = resolved
                        answer_val = resolved or str(raw).strip()
                        upsert_answer(cur, rid, fc, answer_val, idx)

                    elif fc in FCODE_CITY:
                        if is_null(raw):
                            continue   # f5a2 kosong → skip total, tidak isi 0
                        resolved = resolve_city_id(raw, province_id_str)
                        answer_val = resolved or str(raw).strip()
                        upsert_answer(cur, rid, fc, answer_val, idx)

                    elif fc == "f505":
                        if is_null(raw):
                            upsert_answer(cur, rid, fc, None, idx)
                        else:
                            salary_clean = clean_salary(raw)
                            answer_val = str(int(salary_clean)) if salary_clean else str(raw).strip()
                            upsert_answer(cur, rid, fc, answer_val, idx)

                    else:
                        if is_null(raw):
                            upsert_answer(cur, rid, fc, None, idx)
                        else:
                            upsert_answer(cur, rid, fc, str(raw).strip(), idx)

                conn.commit()
                s_ok += 1
                ok   += 1

            except Exception as e:
                conn.rollback()
                s_err += 1
                err   += 1
                print(f"    [ERROR] {sname} baris {rnum} ({nama}): {e}")
                if s_err <= 3:
                    traceback.print_exc()

        print(f"  ✓ {sname}: {s_ok} OK, {s_err} error")

    cur.close()
    print(f"\n  FILE TOTAL: {ok} berhasil, {err} error")


def main():
    print("=" * 60)
    print("SmartTracer ETL v5")
    print("=" * 60)

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.autocommit = False
        print(f"✓ Konek ke: {DB_CONFIG['dbname']}")
    except Exception as e:
        print(f"✗ Gagal konek: {e}")
        sys.exit(1)

    cur = conn.cursor()
    load_geo(cur)
    cur.close()

    if not RUN_ALL:
        print(f"\n[TEST] {TEST_FILE} | grad_year={TEST_GRAD_YEAR} | "
              f"max_rows={TEST_MAX_ROWS if 'TEST_MAX_ROWS' in dir() else 'semua'}")
        process_file(TEST_FILE, TEST_GRAD_YEAR, conn,
                     TEST_MAX_ROWS if 'TEST_MAX_ROWS' in dir() else None)
    else:
        for f, y in EXCEL_FILES:
            if not os.path.exists(f):
                print(f"\n[SKIP] File tidak ada: {f}")
                continue
            print(f"\n{'='*50}\n{f} | tahun lulus: {y}\n{'='*50}")
            process_file(f, y, conn)

    conn.close()
    print("\n✓ Selesai!")
    print("""
Verifikasi di pgAdmin:

-- 1. Cek f502 sudah tidak ada duplikat (semua answer_index = 0)
SELECT answer_index, COUNT(*) FROM tracer_oltp.response_answers
WHERE question_code = 'f502' GROUP BY answer_index ORDER BY answer_index;

-- 2. Cek null sudah terisi "0"
SELECT question_code, COUNT(*) FILTER (WHERE answer_text = '0') AS jumlah_nol,
       COUNT(*) AS total
FROM tracer_oltp.response_answers
GROUP BY question_code ORDER BY question_code;

-- 3. Jumlah jawaban per question_code (kelengkapan data)
SELECT question_code, COUNT(*) AS jumlah
FROM tracer_oltp.response_answers
GROUP BY question_code ORDER BY question_code;

-- 4. Pastikan tidak ada duplikat response per alumni
SELECT response_count, COUNT(*) as jumlah_alumni
FROM (
    SELECT alumni_id, COUNT(*) as response_count
    FROM tracer_oltp.responses GROUP BY alumni_id
) t GROUP BY response_count ORDER BY response_count;
""")


if __name__ == "__main__":
    main()